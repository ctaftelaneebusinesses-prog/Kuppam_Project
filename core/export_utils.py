"""
Super Admin data export pipeline — streams the Users and Posts dashboard
tables out as .xlsx (openpyxl, mirroring excel_utils.build_sample_workbook's
style) or .pdf (reportlab, pure-Python so it needs no system libraries).
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet


def _autosize(ws):
    for column_cells in ws.columns:
        length = max((len(str(c.value)) for c in column_cells if c.value is not None), default=0)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 4, 10), 45)


def _build_workbook(title, headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    ws.append(headers)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for row in rows:
        ws.append(row)

    _autosize(ws)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _build_pdf(title, headers, rows):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), title=title,
                             leftMargin=0.4 * inch, rightMargin=0.4 * inch,
                             topMargin=0.4 * inch, bottomMargin=0.4 * inch)
    styles = getSampleStyleSheet()
    story = [Paragraph(title, styles['Title'])]

    data = [headers] + [[str(cell) for cell in row] for row in rows]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0D6EFD')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    story.append(table)
    doc.build(story)
    buffer.seek(0)
    return buffer


USER_HEADERS = ['Name', 'Email', 'Username', 'Role', 'City', 'Status', 'Joined']


def _user_row(profile):
    status = 'Blocked' if profile.is_blocked else ('Suspended' if profile.is_suspended else 'Active')
    return [
        profile.full_name or profile.user.username,
        profile.user.email,
        profile.user.username,
        profile.get_role_display(),
        profile.city or '—',
        status,
        profile.created_at.strftime('%d %b %Y'),
    ]


def build_users_workbook(profiles_qs):
    return _build_workbook('Users', USER_HEADERS, [_user_row(p) for p in profiles_qs])


def build_users_pdf(profiles_qs):
    return _build_pdf('OneTownCity — Users', USER_HEADERS, [_user_row(p) for p in profiles_qs])


POST_HEADERS = ['Title', 'Type', 'Category', 'Owner', 'Status', 'Views', 'Likes', 'Created']


def _post_row(item):
    obj = item['obj']
    return [
        str(obj),
        item['model_key'].title(),
        obj.listing_category.label if obj.listing_category else '—',
        (obj.owner.profile.full_name or obj.owner.email) if obj.owner else 'Staff',
        obj.get_status_display(),
        obj.view_count,
        obj.like_count,
        obj.created_at.strftime('%d %b %Y'),
    ]


def build_posts_workbook(items):
    return _build_workbook('Posts', POST_HEADERS, [_post_row(i) for i in items])


def build_posts_pdf(items):
    return _build_pdf('OneTownCity — Posts', POST_HEADERS, [_post_row(i) for i in items])
