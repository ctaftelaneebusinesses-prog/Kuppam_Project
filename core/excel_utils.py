"""
Excel bulk-upload engine for OneTownCity.

Reads .xlsx/.xls files with pandas (openpyxl engine), validates each row,
and inserts or updates records directly in the Supabase PostgreSQL database
via the Django ORM. Also builds downloadable sample templates with openpyxl.
"""
import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from dateutil import parser as date_parser

from django.core.exceptions import MultipleObjectsReturned, ValidationError as DjangoValidationError
from django.core.validators import URLValidator
from django.db import DataError, IntegrityError, transaction
from django.utils import timezone

from .models import Business, Property, Job, Event, News, Project


class RowValidationError(Exception):
    """Raised when a single Excel row fails validation. Caught per-row so
    one bad row doesn't stop the rest of the file from being processed."""
    pass


class ExcelValidationError(Exception):
    """Raised when the file itself (not a single row) can't be processed —
    e.g. unreadable file or missing required columns."""
    pass


@dataclass
class UploadResult:
    total: int
    inserted: int
    updated: int
    errors: list = field(default_factory=list)

    @property
    def has_errors(self):
        return bool(self.errors)

    @property
    def success_count(self):
        return self.inserted + self.updated


# ------------------------------------------------------------------
# Cell-level helpers
# ------------------------------------------------------------------
def cell_to_str(value):
    """Normalize a raw pandas/openpyxl cell value to a clean string."""
    if value is None:
        return ''
    if isinstance(value, float):
        if pd.isna(value):
            return ''
        if value.is_integer():
            return str(int(value))
        return str(value)
    if isinstance(value, str):
        if value.strip().lower() == 'nan':
            return ''
        return value.strip()
    return str(value).strip()


def clean_str(value):
    return cell_to_str(value)


def parse_decimal(value, field_name):
    text = cell_to_str(value).replace(',', '').replace('₹', '').replace('Rs.', '').replace('Rs', '').strip()
    if not text:
        raise RowValidationError(f'{field_name} is required')
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        raise RowValidationError(f'{field_name} "{value}" is not a valid number')


def match_choice(value, choices, field_name):
    value_norm = cell_to_str(value).strip().lower()
    for key, label in choices:
        if value_norm == key.lower() or value_norm == label.lower():
            return key
    valid = ', '.join(label for _, label in choices)
    raise RowValidationError(f'Invalid {field_name} "{value}". Must be one of: {valid}')


def validate_url(value, max_length=500):
    text = cell_to_str(value)
    if not text:
        return ''
    validator = URLValidator()
    try:
        validator(text)
    except DjangoValidationError:
        raise RowValidationError(f'"{text}" is not a valid URL')
    if len(text) > max_length:
        raise RowValidationError(
            f'Image URL is too long ({len(text)} characters, max {max_length}); '
            'please use a shorter direct image link'
        )
    return text


def clean_phone(value, field_name='Contact', required=True):
    text = cell_to_str(value)
    if not text:
        if required:
            raise RowValidationError(f'{field_name} is required')
        return ''
    cleaned = re.sub(r'[^\d+]', '', text)
    digits_only = re.sub(r'\D', '', cleaned)
    if len(digits_only) < 10:
        raise RowValidationError(f'{field_name} "{value}" must contain at least 10 digits')
    return cleaned[:15]


def parse_bool(value, default=True):
    text = cell_to_str(value).lower()
    if not text:
        return default
    if text in ('yes', 'y', 'true', '1'):
        return True
    if text in ('no', 'n', 'false', '0'):
        return False
    return default


def parse_date(value, field_name, required=True):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        if required:
            raise RowValidationError(f'{field_name} is required')
        return None
    if hasattr(value, 'date') and callable(getattr(value, 'date')):
        # pandas.Timestamp / datetime.datetime
        try:
            if pd.isna(value):
                if required:
                    raise RowValidationError(f'{field_name} is required')
                return None
        except TypeError:
            pass
        return value.date()
    import datetime as _dt
    if isinstance(value, _dt.date):
        return value

    text = cell_to_str(value)
    if not text or text.lower() == 'nat':
        if required:
            raise RowValidationError(f'{field_name} is required')
        return None
    try:
        return date_parser.parse(text, dayfirst=True).date()
    except (ValueError, OverflowError):
        raise RowValidationError(f'{field_name} "{value}" is not a valid date (use YYYY-MM-DD)')


def cell_is_empty(value):
    if value is None:
        return True
    if isinstance(value, float) and pd.isna(value):
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


# ------------------------------------------------------------------
# Per-model row parsers.
# Each returns (lookup_kwargs, defaults_kwargs) used for update_or_create,
# or raises RowValidationError with a human-readable message.
# ------------------------------------------------------------------
def parse_business_row(row):
    name = clean_str(row.get('Business Name'))
    if not name:
        raise RowValidationError('Business Name is required')
    address = clean_str(row.get('Address'))
    if not address:
        raise RowValidationError('Address is required')
    phone = clean_phone(row.get('Phone Number'), 'Phone Number')
    category_raw = clean_str(row.get('Category'))
    category = match_choice(category_raw, Business.CATEGORY_CHOICES, 'Category') if category_raw else 'other'

    lookup = {'name': name, 'phone_number': phone}
    defaults = {
        'category': category,
        'address': address,
        'image_url': validate_url(row.get('Image URL')),
        'description': clean_str(row.get('Description')),
        'is_featured': parse_bool(row.get('Featured'), default=False),
        'is_active': parse_bool(row.get('Active'), default=True),
    }
    return lookup, defaults


def make_directory_row_parser(category_value):
    """
    Builds a row parser for a Business-backed directory category (Restaurants,
    Hospitals, Transport) with the category fixed — so the Excel file for
    these doesn't need a Category column at all.
    """
    def parse_row(row):
        name = clean_str(row.get('Name'))
        if not name:
            raise RowValidationError('Name is required')
        address = clean_str(row.get('Address'))
        if not address:
            raise RowValidationError('Address is required')
        phone = clean_phone(row.get('Phone Number'), 'Phone Number')

        lookup = {'name': name, 'phone_number': phone}
        defaults = {
            'category': category_value,
            'address': address,
            'image_url': validate_url(row.get('Image URL')),
            'description': clean_str(row.get('Description')),
            'is_featured': parse_bool(row.get('Featured'), default=False),
            'is_active': parse_bool(row.get('Active'), default=True),
        }
        return lookup, defaults
    return parse_row


def make_subcategory_row_parser(subcategory_choices, default_category):
    """
    Builds a row parser for a directory category that groups more than one
    Business.category value under one page (e.g. Education = School +
    College & University). The Excel file gets a 'Category' column so each
    row is tagged with its specific sub-category; blank cells fall back to
    default_category.
    """
    def parse_row(row):
        name = clean_str(row.get('Name'))
        if not name:
            raise RowValidationError('Name is required')
        address = clean_str(row.get('Address'))
        if not address:
            raise RowValidationError('Address is required')
        phone = clean_phone(row.get('Phone Number'), 'Phone Number')
        category_raw = clean_str(row.get('Category'))
        category = match_choice(category_raw, subcategory_choices, 'Category') if category_raw else default_category

        lookup = {'name': name, 'phone_number': phone}
        defaults = {
            'category': category,
            'address': address,
            'image_url': validate_url(row.get('Image URL')),
            'description': clean_str(row.get('Description')),
            'is_featured': parse_bool(row.get('Featured'), default=False),
            'is_active': parse_bool(row.get('Active'), default=True),
        }
        return lookup, defaults
    return parse_row


def parse_property_row(row):
    title = clean_str(row.get('Title'))
    if not title:
        raise RowValidationError('Title is required')
    location = clean_str(row.get('Location'))
    if not location:
        raise RowValidationError('Location is required')
    type_raw = clean_str(row.get('Type'))
    property_type = match_choice(type_raw, Property.PROPERTY_TYPE_CHOICES, 'Type') if type_raw else 'sale'
    price = parse_decimal(row.get('Price'), 'Price')
    contact = clean_phone(row.get('Contact'), 'Contact')

    lookup = {'title': title, 'location': location}
    defaults = {
        'property_type': property_type,
        'price': price,
        'contact_number': contact,
        'image_url': validate_url(row.get('Image URL')),
        'description': clean_str(row.get('Description')),
        'is_featured': parse_bool(row.get('Featured'), default=False),
        'is_active': parse_bool(row.get('Active'), default=True),
    }
    return lookup, defaults


def parse_job_row(row):
    company = clean_str(row.get('Company'))
    if not company:
        raise RowValidationError('Company is required')
    job_title = clean_str(row.get('Job Title'))
    if not job_title:
        raise RowValidationError('Job Title is required')
    location = clean_str(row.get('Location'))
    if not location:
        raise RowValidationError('Location is required')
    contact = clean_phone(row.get('Contact'), 'Contact')

    lookup = {'job_title': job_title, 'company': company}
    defaults = {
        'location': location,
        'salary': clean_str(row.get('Salary')),
        'contact_number': contact,
        'description': clean_str(row.get('Description')),
        'is_featured': parse_bool(row.get('Featured'), default=False),
        'is_active': parse_bool(row.get('Active'), default=True),
    }
    return lookup, defaults


def parse_event_row(row):
    title = clean_str(row.get('Title'))
    if not title:
        raise RowValidationError('Title is required')
    event_date = parse_date(row.get('Event Date'), 'Event Date', required=True)
    location = clean_str(row.get('Location'))
    if not location:
        raise RowValidationError('Location is required')

    lookup = {'title': title, 'event_date': event_date}
    defaults = {
        'location': location,
        'description': clean_str(row.get('Description')),
        'contact_number': clean_phone(row.get('Contact'), 'Contact', required=False),
        'image_url': validate_url(row.get('Image URL')),
        'is_featured': parse_bool(row.get('Featured'), default=False),
        'is_active': parse_bool(row.get('Active'), default=True),
    }
    return lookup, defaults


def parse_news_row(row):
    title = clean_str(row.get('Title'))
    if not title:
        raise RowValidationError('Title is required')
    content = clean_str(row.get('Content'))
    if not content:
        raise RowValidationError('Content is required')
    published_date = parse_date(row.get('Published Date'), 'Published Date', required=False) or timezone.localdate()

    lookup = {'title': title, 'published_date': published_date}
    defaults = {
        'content': content,
        'image_url': validate_url(row.get('Image URL')),
        'source': clean_str(row.get('Source')),
        'is_featured': parse_bool(row.get('Featured'), default=False),
        'is_active': parse_bool(row.get('Active'), default=True),
    }
    return lookup, defaults


def parse_project_row(row):
    title = clean_str(row.get('Title'))
    if not title:
        raise RowValidationError('Title is required')
    location = clean_str(row.get('Location'))
    if not location:
        raise RowValidationError('Location is required')
    status_raw = clean_str(row.get('Status'))
    project_status = match_choice(status_raw, Project.STATUS_CHOICES, 'Status') if status_raw else 'planned'

    lookup = {'title': title, 'location': location}
    defaults = {
        'project_status': project_status,
        'expected_completion': parse_date(row.get('Expected Completion'), 'Expected Completion', required=False),
        'department': clean_str(row.get('Department')),
        'description': clean_str(row.get('Description')),
        'image_url': validate_url(row.get('Image URL')),
        'is_featured': parse_bool(row.get('Featured'), default=False),
        'is_active': parse_bool(row.get('Active'), default=True),
    }
    return lookup, defaults


# ------------------------------------------------------------------
# Upload configuration registry — one entry per module.
# ------------------------------------------------------------------
UPLOAD_CONFIGS = {
    'businesses': {
        'key': 'businesses',
        'label': 'Businesses',
        'icon': 'bi-shop',
        'model': Business,
        'required_columns': ['Business Name', 'Category', 'Address', 'Phone Number'],
        'optional_columns': ['Image URL', 'Description', 'Featured', 'Active'],
        'parse_row': parse_business_row,
        'list_url_name': 'core:business_list',
        'example_row': {
            'Business Name': 'Sri Lakshmi Grocery',
            'Category': 'Grocery Store',
            'Address': 'Main Bazaar Road, Kuppam',
            'Phone Number': '9876543210',
            'Image URL': 'https://example.com/images/shop.jpg',
            'Description': 'Daily essentials, groceries and household items.',
            'Featured': 'No',
            'Active': 'Yes',
        },
    },
    'restaurants': {
        'key': 'restaurants',
        'label': 'Restaurants',
        'icon': 'bi-cup-hot',
        'model': Business,
        'required_columns': ['Name', 'Address', 'Phone Number'],
        'optional_columns': ['Image URL', 'Description', 'Featured', 'Active'],
        'parse_row': make_directory_row_parser('restaurant'),
        'list_url_name': 'core:restaurant_list',
        'example_row': {
            'Name': 'Kuppam Tiffin Centre',
            'Address': 'Market Street, Kuppam',
            'Phone Number': '9876543210',
            'Image URL': 'https://example.com/images/restaurant.jpg',
            'Description': 'South Indian breakfast and tiffins.',
            'Featured': 'No',
            'Active': 'Yes',
        },
    },
    'hospitals': {
        'key': 'hospitals',
        'label': 'Hospitals & Healthcare',
        'icon': 'bi-hospital',
        'model': Business,
        'required_columns': ['Name', 'Address', 'Phone Number'],
        'optional_columns': ['Image URL', 'Description', 'Featured', 'Active'],
        'parse_row': make_directory_row_parser('hospital'),
        'list_url_name': 'core:hospital_list',
        'example_row': {
            'Name': 'Kuppam General Hospital',
            'Address': 'Hospital Road, Kuppam',
            'Phone Number': '9876543210',
            'Image URL': 'https://example.com/images/hospital.jpg',
            'Description': '24x7 emergency and outpatient care.',
            'Featured': 'No',
            'Active': 'Yes',
        },
    },
    'education': {
        'key': 'education',
        'label': 'Education',
        'icon': 'bi-mortarboard',
        'model': Business,
        'required_columns': ['Name', 'Address', 'Phone Number'],
        'optional_columns': ['Category', 'Image URL', 'Description', 'Featured', 'Active'],
        'parse_row': make_subcategory_row_parser(
            [('school', 'School'), ('college', 'College & University')], default_category='school',
        ),
        'list_url_name': 'core:education_list',
        'example_row': {
            'Name': 'Kuppam Public School',
            'Category': 'School',
            'Address': 'School Road, Kuppam',
            'Phone Number': '9876543210',
            'Image URL': 'https://example.com/images/school.jpg',
            'Description': 'CBSE school, classes 1 to 10.',
            'Featured': 'No',
            'Active': 'Yes',
        },
    },
    'transport': {
        'key': 'transport',
        'label': 'Transport',
        'icon': 'bi-bus-front',
        'model': Business,
        'required_columns': ['Name', 'Address', 'Phone Number'],
        'optional_columns': ['Image URL', 'Description', 'Featured', 'Active'],
        'parse_row': make_directory_row_parser('transport'),
        'list_url_name': 'core:transport_list',
        'example_row': {
            'Name': 'Kuppam Travels',
            'Address': 'Bus Stand Road, Kuppam',
            'Phone Number': '9876543210',
            'Image URL': 'https://example.com/images/transport.jpg',
            'Description': 'Taxi and tempo travel booking.',
            'Featured': 'No',
            'Active': 'Yes',
        },
    },
    'properties': {
        'key': 'properties',
        'label': 'Properties',
        'icon': 'bi-house-door',
        'model': Property,
        'required_columns': ['Title', 'Type', 'Price', 'Location', 'Contact'],
        'optional_columns': ['Image URL', 'Description', 'Featured', 'Active'],
        'parse_row': parse_property_row,
        'list_url_name': 'core:property_list',
        'example_row': {
            'Title': '2BHK Flat near Bus Stand',
            'Type': 'For Rent',
            'Price': '15000',
            'Location': 'RTC Bus Stand Road, Kuppam',
            'Contact': '9876543210',
            'Image URL': 'https://example.com/images/flat.jpg',
            'Description': 'Spacious 2BHK with covered parking.',
            'Featured': 'No',
            'Active': 'Yes',
        },
    },
    'jobs': {
        'key': 'jobs',
        'label': 'Jobs',
        'icon': 'bi-briefcase',
        'model': Job,
        'required_columns': ['Company', 'Job Title', 'Location', 'Contact'],
        'optional_columns': ['Salary', 'Description', 'Featured', 'Active'],
        'parse_row': parse_job_row,
        'list_url_name': 'core:job_list',
        'example_row': {
            'Company': 'Kuppam Textiles',
            'Job Title': 'Sales Executive',
            'Location': 'Kuppam',
            'Salary': '₹12,000 - ₹18,000/month',
            'Contact': '9876543210',
            'Description': 'Looking for an energetic sales executive.',
            'Featured': 'No',
            'Active': 'Yes',
        },
    },
    'events': {
        'key': 'events',
        'label': 'Events',
        'icon': 'bi-calendar-event',
        'model': Event,
        'required_columns': ['Title', 'Event Date', 'Location'],
        'optional_columns': ['Description', 'Contact', 'Image URL', 'Featured', 'Active'],
        'parse_row': parse_event_row,
        'list_url_name': 'core:event_list',
        'example_row': {
            'Title': 'Kuppam Cultural Festival',
            'Event Date': '2026-08-15',
            'Location': 'Town Hall Grounds, Kuppam',
            'Description': 'Annual cultural festival with music and food stalls.',
            'Contact': '9876543210',
            'Image URL': 'https://example.com/images/event.jpg',
            'Featured': 'No',
            'Active': 'Yes',
        },
    },
    'news': {
        'key': 'news',
        'label': 'News',
        'icon': 'bi-newspaper',
        'model': News,
        'required_columns': ['Title', 'Content'],
        'optional_columns': ['Published Date', 'Image URL', 'Source', 'Featured', 'Active'],
        'parse_row': parse_news_row,
        'list_url_name': 'core:news_list',
        'example_row': {
            'Title': 'New Water Supply Scheme Launched',
            'Content': 'The municipality has launched a new water supply scheme for Kuppam residents.',
            'Published Date': '2026-07-20',
            'Image URL': 'https://example.com/images/news.jpg',
            'Source': 'Kuppam Municipality',
            'Featured': 'No',
            'Active': 'Yes',
        },
    },
    'projects': {
        'key': 'projects',
        'label': 'Upcoming Projects',
        'icon': 'bi-cone-striped',
        'model': Project,
        'required_columns': ['Title', 'Location'],
        'optional_columns': ['Status', 'Expected Completion', 'Department', 'Description', 'Image URL', 'Featured', 'Active'],
        'parse_row': parse_project_row,
        'list_url_name': 'core:project_list',
        'example_row': {
            'Title': 'Swarna Kuppam Road Upgrade',
            'Location': 'Internal roads, Kuppam',
            'Status': 'Ongoing',
            'Expected Completion': '2027-03-31',
            'Department': 'Kuppam Municipality',
            'Description': 'Upgrading internal roads across the constituency under the Swarna Kuppam initiative.',
            'Image URL': 'https://example.com/images/project.jpg',
            'Featured': 'No',
            'Active': 'Yes',
        },
    },
}


# ------------------------------------------------------------------
# Core engine
# ------------------------------------------------------------------
def process_excel_upload(uploaded_file, config):
    """
    Reads the uploaded Excel file, validates its columns, and inserts or
    updates records for the given module config. Returns an UploadResult.
    Raises ExcelValidationError for file-level problems (unreadable file,
    missing required columns).
    """
    try:
        df = pd.read_excel(uploaded_file, engine='openpyxl')
    except Exception as exc:
        raise ExcelValidationError(f'Could not read the Excel file: {exc}')

    df.columns = [str(c).strip() for c in df.columns]

    missing = [col for col in config['required_columns'] if col not in df.columns]
    if missing:
        raise ExcelValidationError(
            'Missing required column(s): ' + ', '.join(missing)
            + '. Download the sample template to see the expected format.'
        )

    if df.empty:
        return UploadResult(total=0, inserted=0, updated=0, errors=[])

    model = config['model']
    parse_row = config['parse_row']

    inserted = 0
    updated = 0
    errors = []

    for idx, row in df.iterrows():
        row_number = idx + 2  # account for the header row
        row_dict = row.to_dict()

        if all(cell_is_empty(v) for v in row_dict.values()):
            continue

        try:
            lookup, defaults = parse_row(row_dict)
        except RowValidationError as exc:
            errors.append(f'Row {row_number}: {exc}')
            continue

        try:
            with transaction.atomic():
                obj, created = model.objects.update_or_create(**lookup, defaults=defaults)
        except MultipleObjectsReturned:
            errors.append(
                f'Row {row_number}: multiple existing records match this row; '
                'please resolve the duplicates in the admin panel first.'
            )
            continue
        except IntegrityError as exc:
            errors.append(f'Row {row_number}: database error - {exc}')
            continue
        except DataError as exc:
            errors.append(f'Row {row_number}: a field value is too long - {exc}')
            continue

        if created:
            inserted += 1
        else:
            updated += 1

    return UploadResult(total=len(df), inserted=inserted, updated=updated, errors=errors)


def build_sample_workbook(config):
    """Builds an in-memory openpyxl Workbook with headers + one example row."""
    headers = list(config['required_columns']) + list(config['optional_columns'])
    example = config.get('example_row', {})

    wb = Workbook()
    ws = wb.active
    ws.title = config['label'][:31]

    ws.append(headers)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    ws.append([example.get(h, '') for h in headers])

    for column_cells in ws.columns:
        length = max((len(str(c.value)) for c in column_cells if c.value is not None), default=0)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 4, 14), 45)

    return wb
